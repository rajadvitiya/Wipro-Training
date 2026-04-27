import openpyxl
import os

# Step 1: Create a new Excel workbook and write data
filename = "example.xlsx"

# Create a workbook and select the active sheet
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "People"

# Write header
sheet.append(["Name", "Age", "City"])

# Write some rows
sheet.append(["Alice", 30, "New York"])
sheet.append(["Bob", 25, "London"])
sheet.append(["Charlie", 35, "Paris"])

# Save the workbook
wb.save(filename)
print(f"{filename} created and data written.")

# Step 2: Read from the Excel file
wb2 = openpyxl.load_workbook(filename)
sheet2 = wb2["People"]

print("Reading data from Excel:")
for row in sheet2.iter_rows(values_only=True):
    print(row)

# Step 3: Delete the Excel file
if os.path.exists(filename):
    os.remove(filename)
    print(f"{filename} has been deleted.")
else:
    print("File does not exist.")
